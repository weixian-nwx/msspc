"""Reading the expected-participants workbook and writing the attendance export.

The output attendance file preserves every original column and the original row
order, simply appending a ``Status`` column (Present / Absent).
"""
from __future__ import annotations

import re
import shutil
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app import config
from app.db import Database


class ExcelError(Exception):
    """Raised when the uploaded workbook cannot be used."""


def _normalize(text: object) -> str:
    """Normalize a header for matching: case-, space- and underscore-insensitive.

    This lets headers like ``Seat No``, ``seat_no`` and ``SEAT NO`` all match the
    same required column key.
    """
    if text is None:
        return ""
    return re.sub(r"[\s_]+", "", str(text).strip().lower())


def _find_required_columns(header: list[object]) -> dict[str, int]:
    """Map each required column name to its 0-based index, tolerating case/space."""
    norm = {_normalize(h): i for i, h in enumerate(header) if h is not None}
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for col in config.REQUIRED_COLUMNS:
        key = _normalize(col)
        if key in norm:
            resolved[col] = norm[key]
        else:
            missing.append(col)
    if missing:
        raise ExcelError(
            "The excel sheet is missing required column(s): "
            + ", ".join(missing)
            + ".\nFound columns: "
            + ", ".join(str(h) for h in header if h is not None)
        )
    return resolved


def import_participants(src_path: str, db: Database) -> int:
    """Parse ``src_path``, copy it into the data folder, load rows into the db.

    Returns the number of participants imported. Raises ExcelError on problems.
    """
    try:
        wb = load_workbook(src_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelError(f"Could not open the excel file:\n{exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ExcelError("The excel sheet is empty.")

    header = list(rows[0])
    cols = _find_required_columns(header)

    parsed: list[dict] = []
    seen_ids: set[str] = set()
    row_index = 0
    for raw in rows[1:]:
        qr = raw[cols[config.COL_QR_ID]] if cols[config.COL_QR_ID] < len(raw) else None
        name = raw[cols[config.COL_NAME]] if cols[config.COL_NAME] < len(raw) else None
        title = raw[cols[config.COL_TITLE]] if cols[config.COL_TITLE] < len(raw) else None
        grade = raw[cols[config.COL_GRADE]] if cols[config.COL_GRADE] < len(raw) else None
        seat = raw[cols[config.COL_SEAT]] if cols[config.COL_SEAT] < len(raw) else None
        bu = raw[cols[config.COL_BU]] if cols[config.COL_BU] < len(raw) else None

        # Skip fully blank rows.
        if all(v is None or str(v).strip() == "" for v in (qr, name, title, grade, seat, bu)):
            continue

        qr_s = "" if qr is None else str(qr).strip()
        if not qr_s:
            raise ExcelError(f"A participant row is missing a unique qr id (row {row_index + 2}).")
        if qr_s in seen_ids:
            raise ExcelError(f"Duplicate unique qr id found: '{qr_s}'. IDs must be unique.")
        seen_ids.add(qr_s)

        parsed.append(
            {
                "qr_id": qr_s,
                "name": "" if name is None else str(name).strip(),
                "title": "" if title is None else str(title).strip(),
                "grade": "" if grade is None else str(grade).strip(),
                "seat_no": "" if seat is None else str(seat).strip(),
                "bu": "" if bu is None else str(bu).strip(),
                "row_index": row_index,
            }
        )
        row_index += 1

    if not parsed:
        raise ExcelError("No participant rows found below the header.")

    # Copy the original file into the data folder (self-contained + crash-safe),
    # then load into the database.
    shutil.copyfile(src_path, config.EXPECTED_XLSX)
    db.replace_participants(parsed)
    db.set_meta("expected_xlsx", config.EXPECTED_XLSX)
    return len(parsed)


def export_attendance(db: Database, out_path: Optional[str] = None) -> str:
    """Write the attendance workbook from the copied original + current status.

    Returns the path written. Re-reads the copied expected workbook so all
    original columns and ordering are preserved, appending a Status column.
    """
    out_path = out_path or config.ATTENDANCE_XLSX
    src = db.get_meta("expected_xlsx")
    if not src:
        raise ExcelError("No expected-participants workbook has been imported yet.")

    wb = load_workbook(src, data_only=True)
    ws = wb.active

    header = [c.value for c in ws[1]]
    norm = {_normalize(h): i for i, h in enumerate(header) if h is not None}
    qr_col = norm[_normalize(config.COL_QR_ID)]  # 0-based

    status_by_id = {p.qr_id: ("Present" if p.present else "Absent") for p in db.all_participants()}

    status_col_idx = len(header)  # 0-based index for the new column
    status_cell = ws.cell(row=1, column=status_col_idx + 1)
    status_cell.value = "Status"

    for r in range(2, ws.max_row + 1):
        qr_val = ws.cell(row=r, column=qr_col + 1).value
        qr_s = "" if qr_val is None else str(qr_val).strip()
        if not qr_s:
            continue
        ws.cell(row=r, column=status_col_idx + 1).value = status_by_id.get(qr_s, "Absent")

    # Light cosmetic touch: widen the Status column.
    ws.column_dimensions[get_column_letter(status_col_idx + 1)].width = 12

    wb.save(out_path)
    wb.close()
    return out_path
